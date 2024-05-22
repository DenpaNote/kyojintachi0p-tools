import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment

cell_style = NamedStyle(name="cell_style")
cell_style.alignment = Alignment(wrap_text=True)

def get_filename_without_extension(path):
    return path.split("/")[-1].split(".")[0]

def parse_xml(xml_path):
    tree = ET.parse(xml_path)
    wb = Workbook()
    wb.add_named_style(cell_style)
    ws = wb.active
    ws.append(["element_iter_index", "japanese_text", "translated_text"])
    for index, child in enumerate(tree.getroot().iter()):
        if child.tail and child.tail.strip():
            ws.append([index, child.tail, ""])
    for row in ws.iter_rows():
        for cell in row:
            cell.style = cell_style

    ws.column_dimensions["B"].width = 200
    ws.column_dimensions["C"].width = 200
    
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top")
    
    wb.save(get_filename_without_extension(xml_path) + ".xlsx")

import os

for file in os.listdir("."):
    if file.endswith(".xml"):
        parse_xml(file)
